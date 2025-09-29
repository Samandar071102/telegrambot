from openpyxl import Workbook, load_workbook
import os

def save_user(user_id, first_name, last_name, phone, subject, score):
    file = "users.xlsx"

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(["UserID", "Ism", "Familiya", "Telefon", "Fan", "Ball"])
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active
    ws.append([user_id, first_name, last_name, phone, subject, score])
    wb.save(file)
