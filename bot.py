import json
import asyncio
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from openpyxl import Workbook, load_workbook
import os

TOKEN = "8433010343:AAFfenNG_z8bBdXxNjzW0dJhRGkFHsjRbs0"   # BotFather dan olingan token
CHANNEL_ID = "@Aqllilik_balosi"           # Kanal username
ADMIN_ID = 1130602920                # O'zingizning Telegram ID'ingizni yozing

# Excel faylni yaratish
if not os.path.exists("users.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["UserID", "Ism", "Familiya", "Telefon", "Fan", "Ball"])
    wb.save("users.xlsx")
    wb.close()

# Bot va dispatcher
bot = Bot(token=TOKEN)
dp = Dispatcher()

# Ma’lumotlarni vaqtincha saqlash
user_data = {}

# Start komandasi
@dp.message(Command("start"))
async def start(message: types.Message):
    member = await bot.get_chat_member(CHANNEL_ID, message.from_user.id)
    if member.status == "left":
        await message.answer("❗ Botdan foydalanish uchun kanalga obuna bo‘ling", 
                             reply_markup=types.InlineKeyboardMarkup(
                                 inline_keyboard=[
                                     [types.InlineKeyboardButton(text="📢 Kanalga obuna bo‘lish", url=f"https://t.me/{CHANNEL_ID[1:]}")]
                                 ]
                             ))
        return

    user_data[message.from_user.id] = {"score": 0, "current_test": []}

    await message.answer("Assalomu alaykum! Ismingizni kiriting:")
    

# Ism qabul qilish
@dp.message(lambda msg: msg.from_user.id in user_data and "ism" not in user_data[msg.from_user.id])
async def get_name(message: types.Message):
    user_data[message.from_user.id]["ism"] = message.text
    await message.answer("Familiyangizni kiriting:")

# Familiya
@dp.message(lambda msg: "ism" in user_data.get(msg.from_user.id, {}) and "familiya" not in user_data[msg.from_user.id])
async def get_surname(message: types.Message):
    user_data[message.from_user.id]["familiya"] = message.text

    # Telefon tugmasi
    kb = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📱 Telefon raqamni yuborish", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True
    )
    await message.answer("Telefon raqamingizni yuboring:", reply_markup=kb)

# Telefon
@dp.message(lambda msg: "familiya" in user_data.get(msg.from_user.id, {}) and "telefon" not in user_data[msg.from_user.id])
async def get_phone(message: types.Message):
    if message.contact:
        user_data[message.from_user.id]["telefon"] = message.contact.phone_number
    else:
        user_data[message.from_user.id]["telefon"] = message.text

    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📐 Matematika"), KeyboardButton(text="🇬🇧 Ingliz tili")],
            [KeyboardButton(text="📖 Ona tili")]
        ],
        resize_keyboard=True
    )
    await message.answer("Qaysi fanga qiziqasiz?", reply_markup=kb)

# Qiziqish
@dp.message(lambda msg: "telefon" in user_data.get(msg.from_user.id, {}) and "fan" not in user_data[msg.from_user.id])
async def get_subject(message: types.Message):
    fan = message.text
    user_data[message.from_user.id]["fan"] = fan
    user_data[message.from_user.id]["score"] = 0

    await message.answer(f"{fan} bo‘yicha testni boshlaymiz!", reply_markup=ReplyKeyboardRemove())
    await send_test(message.from_user.id, fan)

# Test yuborish
async def send_test(user_id, fan):
    file_map = {
        "📐 Matematika": "tests/matematika.json",
        "🇬🇧 Ingliz tili": "tests/ingliz.json",
        "📖 Ona tili": "tests/ona_tili.json"
    }
    path = file_map.get(fan)
    if not path or not os.path.exists(path):
        await bot.send_message(user_id, "❌ Bu fan bo‘yicha test topilmadi.")
        return

    with open(path, "r", encoding="utf-8") as f:
        savollar = json.load(f)

    user_data[user_id]["current_test"] = savollar
    user_data[user_id]["index"] = 0
    await ask_question(user_id)

async def ask_question(user_id):
    index = user_data[user_id]["index"]
    savollar = user_data[user_id]["current_test"]

    if index >= len(savollar):
        score = user_data[user_id]["score"]

        # Excelga yozish
        wb = load_workbook("users.xlsx")
        ws = wb.active
        ws.append([
            user_id,
            user_data[user_id]["ism"],
            user_data[user_id]["familiya"],
            user_data[user_id]["telefon"],
            user_data[user_id]["fan"],
            score
        ])
        wb.save("users.xlsx")
        wb.close()

        await bot.send_message(user_id, f"✅ Test tugadi!\nSizning natijangiz: {score}/{len(savollar)}")
        return

    s = savollar[index]
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(*s["variantlar"])
    await bot.send_message(user_id, s["savol"], reply_markup=kb)

# Javobni tekshirish
@dp.message(lambda msg: "current_test" in user_data.get(msg.from_user.id, {}))
async def check_answer(message: types.Message):
    user_id = message.from_user.id
    index = user_data[user_id]["index"]
    savollar = user_data[user_id]["current_test"]

    if index < len(savollar):
        togrijavob = savollar[index]["javob"]
        if message.text == togrijavob:
            user_data[user_id]["score"] += 1
            await message.answer("✅ To‘g‘ri!")
        else:
            await message.answer(f"❌ Noto‘g‘ri. To‘g‘ri javob: {togrijavob}")

        user_data[user_id]["index"] += 1
        await ask_question(user_id)

# Excel faylni yuklab olish (faqat admin uchun)
@dp.message(Command("download"))
async def download_excel(message: types.Message):
    if message.from_user.id == ADMIN_ID:
        if os.path.exists("users.xlsx"):
            await message.answer_document(types.FSInputFile("users.xlsx"))
        else:
            await message.answer("❌ Fayl topilmadi.")
    else:
        await message.answer("❌ Sizda huquq yo‘q.")

# Run
async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
